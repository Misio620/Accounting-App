"""
記帳應用程式 - 主視窗界面（完整版本）
使用 tkinter 建立桌面應用程式界面
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import sys
import os
import csv
import json
from collections import defaultdict
import calendar

# 圖表相關 imports
try:
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    from matplotlib.figure import Figure
    
    # 設定中文字體
    plt.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei', 'Arial Unicode MS', 'DejaVu Sans']
    plt.rcParams['axes.unicode_minus'] = False
    
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False

# 將專案根目錄加入路徑
sys.path.append(os.path.dirname(os.path.dirname(__file__)))

try:
    from database.models import DatabaseManager, CategoryManager, TransactionManager
except ImportError:
    print("❌ 無法導入資料庫模組，請確認 database/models.py 存在")
    sys.exit(1)

class TransactionDialog:
    """交易記錄新增/編輯對話框"""
    
    def __init__(self, parent, category_manager, transaction_manager, transaction_data=None):
        self.parent = parent
        self.category_manager = category_manager
        self.transaction_manager = transaction_manager
        self.transaction_data = transaction_data
        self.result = None
        
        # 建立對話框視窗
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("新增交易記錄" if transaction_data is None else "編輯交易記錄")
        self.dialog.geometry("400x350")
        self.dialog.transient(parent)
        self.dialog.grab_set()
        self.dialog.resizable(False, False)
        
        self.setup_ui()
        self.center_window()
        
        if transaction_data:
            self.fill_existing_data()
    
    def center_window(self):
        """將對話框置中顯示"""
        self.dialog.update_idletasks()
        x = (self.dialog.winfo_screenwidth() // 2) - (400 // 2)
        y = (self.dialog.winfo_screenheight() // 2) - (350 // 2)
        self.dialog.geometry(f"400x350+{x}+{y}")
    
    def setup_ui(self):
        """設定對話框界面"""
        main_frame = ttk.Frame(self.dialog, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 標題
        title_text = "新增交易記錄" if self.transaction_data is None else "編輯交易記錄"
        ttk.Label(main_frame, text=title_text, font=("Arial", 12, "bold")).grid(
            row=0, column=0, columnspan=2, pady=(0, 15))
        
        # 日期選擇
        ttk.Label(main_frame, text="日期:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.date_var = tk.StringVar(value=datetime.now().strftime('%Y-%m-%d'))
        date_entry = ttk.Entry(main_frame, textvariable=self.date_var, width=15)
        date_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), pady=5)
        
        # 添加日期格式提示
        ttk.Label(main_frame, text="(格式: YYYY-MM-DD)", font=("Arial", 8), 
                 foreground="gray").grid(row=2, column=1, sticky=tk.W)
        
        # 類型選擇
        ttk.Label(main_frame, text="類型:").grid(row=3, column=0, sticky=tk.W, pady=5)
        self.type_var = tk.StringVar(value="expense")
        type_frame = ttk.Frame(main_frame)
        type_frame.grid(row=3, column=1, sticky=(tk.W, tk.E), pady=5)
        
        ttk.Radiobutton(type_frame, text="收入", variable=self.type_var, 
                       value="income", command=self.on_type_change).pack(side=tk.LEFT)
        ttk.Radiobutton(type_frame, text="支出", variable=self.type_var, 
                       value="expense", command=self.on_type_change).pack(side=tk.LEFT, padx=(20, 0))
        
        # 分類選擇
        ttk.Label(main_frame, text="分類:").grid(row=4, column=0, sticky=tk.W, pady=5)
        self.category_var = tk.StringVar()
        self.category_combo = ttk.Combobox(main_frame, textvariable=self.category_var, 
                                         state="readonly", width=25)
        self.category_combo.grid(row=4, column=1, sticky=(tk.W, tk.E), pady=5)
        
        # 金額輸入
        ttk.Label(main_frame, text="金額:").grid(row=5, column=0, sticky=tk.W, pady=5)
        amount_frame = ttk.Frame(main_frame)
        amount_frame.grid(row=5, column=1, sticky=(tk.W, tk.E), pady=5)
        
        ttk.Label(amount_frame, text="$").pack(side=tk.LEFT)
        self.amount_var = tk.StringVar()
        amount_entry = ttk.Entry(amount_frame, textvariable=self.amount_var, width=22)
        amount_entry.pack(side=tk.LEFT, padx=(3, 0))
        
        # 備註輸入
        ttk.Label(main_frame, text="備註:").grid(row=6, column=0, sticky=tk.W, pady=5)
        self.description_var = tk.StringVar()
        description_entry = ttk.Entry(main_frame, textvariable=self.description_var, width=30)
        description_entry.grid(row=6, column=1, sticky=(tk.W, tk.E), pady=5)
        
        # 按鈕區域
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=7, column=0, columnspan=2, pady=20)
        
        ttk.Button(button_frame, text="確定", command=self.on_ok).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="取消", command=self.on_cancel).pack(side=tk.LEFT, padx=5)
        
        # 設定欄位調整
        main_frame.columnconfigure(1, weight=1)
        
        # 初始載入分類
        self.on_type_change()
        
        # 綁定 Enter 鍵
        self.dialog.bind('<Return>', lambda e: self.on_ok())
        self.dialog.bind('<Escape>', lambda e: self.on_cancel())
    
    def on_type_change(self):
        """當類型改變時更新分類選項"""
        transaction_type = self.type_var.get()
        categories = self.category_manager.get_categories_by_type(transaction_type)
        
        category_names = [f"{cat['id']}: {cat['name']}" for cat in categories]
        self.category_combo['values'] = category_names
        
        if category_names:
            self.category_combo.set(category_names[0])
    
    def fill_existing_data(self):
        """填入現有交易資料（編輯模式）"""
        data = self.transaction_data
        self.date_var.set(data['date'])
        self.type_var.set(data['type'])
        self.amount_var.set(str(data['amount']))
        self.description_var.set(data.get('description', ''))
        
        self.on_type_change()
        
        # 設定對應的分類
        categories = self.category_manager.get_categories_by_type(data['type'])
        for cat in categories:
            if cat['name'] == data['category_name']:
                self.category_combo.set(f"{cat['id']}: {cat['name']}")
                break
    
    def validate_input(self):
        """驗證輸入資料"""
        # 驗證日期
        date_str = self.date_var.get().strip()
        if not date_str:
            messagebox.showerror("錯誤", "請輸入日期")
            return False
        
        try:
            datetime.strptime(date_str, '%Y-%m-%d')
        except ValueError:
            messagebox.showerror("錯誤", "日期格式錯誤，請使用 YYYY-MM-DD 格式")
            return False
        
        # 驗證類型
        if not self.type_var.get():
            messagebox.showerror("錯誤", "請選擇類型")
            return False
        
        # 驗證分類
        if not self.category_var.get():
            messagebox.showerror("錯誤", "請選擇分類")
            return False
        
        # 驗證金額
        amount_str = self.amount_var.get().strip()
        if not amount_str:
            messagebox.showerror("錯誤", "請輸入金額")
            return False
        
        try:
            amount = float(amount_str)
            if amount <= 0:
                messagebox.showerror("錯誤", "金額必須大於 0")
                return False
        except ValueError:
            messagebox.showerror("錯誤", "金額格式錯誤")
            return False
        
        return True
    
    def on_ok(self):
        """確定按鈕處理"""
        if not self.validate_input():
            return
        
        try:
            category_id = int(self.category_var.get().split(':')[0])
            
            self.result = {
                'date': self.date_var.get().strip(),
                'type': self.type_var.get(),
                'category_id': category_id,
                'amount': float(self.amount_var.get().strip()),
                'description': self.description_var.get().strip()
            }
            
            self.dialog.destroy()
            
        except Exception as e:
            messagebox.showerror("錯誤", f"資料處理失敗：{str(e)}")
    
    def on_cancel(self):
        """取消按鈕處理"""
        self.result = None
        self.dialog.destroy()

class MainWindow:
    """主視窗類別"""
    
    def __init__(self):
        print("正在初始化個人記帳本...")
        
        # 初始化資料庫
        try:
            self.db_manager = DatabaseManager("accounting.db")
            self.category_manager = CategoryManager(self.db_manager)
            self.transaction_manager = TransactionManager(self.db_manager)
            print("✅ 資料庫初始化完成")
        except Exception as e:
            print(f"❌ 資料庫初始化失敗: {e}")
            messagebox.showerror("資料庫錯誤", f"無法初始化資料庫：{e}")
            sys.exit(1)
        
        # 建立主視窗
        self.root = tk.Tk()
        self.root.title("個人記帳本 v1.0")
        self.root.geometry("1000x800")
        self.root.minsize(800, 600)
        
        # 設定圖示（如果有的話）
        try:
            # self.root.iconbitmap('icon.ico')  # 可以添加圖示檔案
            pass
        except:
            pass
        
        self.current_transactions = []
        
        self.setup_ui()
        self.refresh_data()
        
        print("✅ 界面初始化完成")
    
    def setup_ui(self):
        """設定主界面"""
        # 主框架
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 標題和版本資訊
        header_frame = ttk.Frame(main_frame)
        header_frame.grid(row=0, column=0, columnspan=3, pady=(0, 10), sticky=(tk.W, tk.E))
        
        ttk.Label(header_frame, text="個人記帳本", font=("Arial", 18, "bold")).pack(side=tk.LEFT)
        ttk.Label(header_frame, text="v1.0", font=("Arial", 10), 
                 foreground="gray").pack(side=tk.RIGHT, padx=(0, 10))
        
        # 按鈕區域
        self.setup_buttons(main_frame)
        
        # 篩選區域
        self.setup_filters(main_frame)
        
        # 交易記錄列表
        self.setup_transaction_list(main_frame)
        
        # 統計區域
        self.setup_statistics(main_frame)
        
        # 報表區域
        self.setup_reports(main_frame)
        
        # 設定網格權重
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(3, weight=1)
        main_frame.rowconfigure(5, weight=1)
        
        # 綁定快捷鍵
        self.setup_shortcuts()
    
    def setup_shortcuts(self):
        """設定快捷鍵"""
        self.root.bind('<Control-n>', lambda e: self.add_transaction())
        self.root.bind('<Control-e>', lambda e: self.edit_transaction())
        self.root.bind('<Delete>', lambda e: self.delete_transaction())
        self.root.bind('<F5>', lambda e: self.refresh_data())
        self.root.bind('<Control-s>', lambda e: self.export_to_csv())
    
    def setup_buttons(self, parent):
        """設定按鈕區域"""
        button_frame = ttk.Frame(parent)
        button_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        
        # 左側按鈕
        left_buttons = ttk.Frame(button_frame)
        left_buttons.pack(side=tk.LEFT)
        
        ttk.Button(left_buttons, text="新增交易 (Ctrl+N)", 
                  command=self.add_transaction).pack(side=tk.LEFT, padx=5)
        ttk.Button(left_buttons, text="編輯交易 (Ctrl+E)", 
                  command=self.edit_transaction).pack(side=tk.LEFT, padx=5)
        ttk.Button(left_buttons, text="刪除交易 (Del)", 
                  command=self.delete_transaction).pack(side=tk.LEFT, padx=5)
        
        # 分隔線
        ttk.Separator(left_buttons, orient='vertical').pack(side=tk.LEFT, padx=10, fill=tk.Y)
        
        ttk.Button(left_buttons, text="重新整理 (F5)", 
                  command=self.refresh_data).pack(side=tk.LEFT, padx=5)
        
        # 右側匯出按鈕
        export_frame = ttk.Frame(button_frame)
        export_frame.pack(side=tk.RIGHT)
        
        ttk.Button(export_frame, text="匯出 CSV", 
                  command=self.export_to_csv).pack(side=tk.LEFT, padx=5)
        ttk.Button(export_frame, text="匯出 Excel", 
                  command=self.export_to_excel).pack(side=tk.LEFT, padx=5)
    
    def setup_filters(self, parent):
        """設定篩選區域"""
        # 篩選標題框架
        filter_title_frame = ttk.Frame(parent)
        filter_title_frame.grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(10, 0))
        
        self.filter_expanded = tk.BooleanVar(value=False)
        self.toggle_button = ttk.Button(filter_title_frame, text="▼ 進階篩選", 
                                       command=self.toggle_filter)
        self.toggle_button.pack(side=tk.LEFT)
        
        ttk.Button(filter_title_frame, text="清除篩選", 
                  command=self.clear_filters).pack(side=tk.LEFT, padx=(10, 0))
        
        # 快速篩選按鈕
        quick_filter_frame = ttk.Frame(filter_title_frame)
        quick_filter_frame.pack(side=tk.RIGHT)
        
        ttk.Button(quick_filter_frame, text="本月", 
                  command=self.filter_current_month).pack(side=tk.LEFT, padx=2)
        ttk.Button(quick_filter_frame, text="本週", 
                  command=self.filter_current_week).pack(side=tk.LEFT, padx=2)
        ttk.Button(quick_filter_frame, text="今日", 
                  command=self.filter_today).pack(side=tk.LEFT, padx=2)
        
        # 篩選內容框架
        self.filter_frame = ttk.LabelFrame(parent, text="篩選條件", padding="10")
        
        # 建立篩選控制項
        self.create_filter_controls()
    
    def create_filter_controls(self):
        """建立篩選控制項"""
        # 第一行
        row1 = ttk.Frame(self.filter_frame)
        row1.grid(row=0, column=0, columnspan=4, sticky=(tk.W, tk.E), pady=5)
        
        ttk.Label(row1, text="起始日期:").pack(side=tk.LEFT, padx=(0, 5))
        self.start_date_var = tk.StringVar()
        ttk.Entry(row1, textvariable=self.start_date_var, width=12).pack(side=tk.LEFT, padx=(0, 15))
        
        ttk.Label(row1, text="結束日期:").pack(side=tk.LEFT, padx=(0, 5))
        self.end_date_var = tk.StringVar()
        ttk.Entry(row1, textvariable=self.end_date_var, width=12).pack(side=tk.LEFT, padx=(0, 15))
        
        ttk.Label(row1, text="類型:").pack(side=tk.LEFT, padx=(0, 5))
        self.type_filter_var = tk.StringVar(value="all")
        ttk.Combobox(row1, textvariable=self.type_filter_var, 
                    values=["all", "income", "expense"], state="readonly", width=10).pack(side=tk.LEFT)
        
        # 第二行
        row2 = ttk.Frame(self.filter_frame)
        row2.grid(row=1, column=0, columnspan=4, sticky=(tk.W, tk.E), pady=5)
        
        ttk.Label(row2, text="分類:").pack(side=tk.LEFT, padx=(0, 5))
        self.category_filter_var = tk.StringVar()
        self.category_filter_combo = ttk.Combobox(row2, textvariable=self.category_filter_var, 
                                                 state="readonly", width=15)
        self.category_filter_combo.pack(side=tk.LEFT, padx=(0, 15))
        
        ttk.Label(row2, text="關鍵字:").pack(side=tk.LEFT, padx=(0, 5))
        self.keyword_var = tk.StringVar()
        ttk.Entry(row2, textvariable=self.keyword_var, width=20).pack(side=tk.LEFT, padx=(0, 15))
        
        ttk.Button(row2, text="套用篩選", command=self.apply_filters).pack(side=tk.LEFT, padx=(15, 0))
        
        # 綁定事件
        self.keyword_var.trace('w', self.on_filter_change)
        self.type_filter_var.trace('w', self.on_filter_change)
        self.category_filter_var.trace('w', self.on_filter_change)
    
    def setup_transaction_list(self, parent):
        """設定交易記錄列表"""
        list_frame = ttk.LabelFrame(parent, text="交易記錄", padding="5")
        list_frame.grid(row=3, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=10)
        
        columns = ('日期', '類型', '分類', '金額', '備註')
        self.transaction_tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=12)
        
        # 設定欄位標題和寬度
        for col in columns:
            self.transaction_tree.heading(col, text=col, command=lambda c=col: self.sort_by_column(c))
        
        self.transaction_tree.column('日期', width=100, anchor='center')
        self.transaction_tree.column('類型', width=80, anchor='center')
        self.transaction_tree.column('分類', width=120, anchor='center')
        self.transaction_tree.column('金額', width=120, anchor='e')
        self.transaction_tree.column('備註', width=250)
        
        # 滾動條
        v_scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.transaction_tree.yview)
        h_scrollbar = ttk.Scrollbar(list_frame, orient=tk.HORIZONTAL, command=self.transaction_tree.xview)
        self.transaction_tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # 放置組件
        self.transaction_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        v_scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        h_scrollbar.grid(row=1, column=0, sticky=(tk.W, tk.E))
        
        # 設定網格權重
        list_frame.columnconfigure(0, weight=1)
        list_frame.rowconfigure(0, weight=1)
        
        # 綁定雙擊事件
        self.transaction_tree.bind('<Double-1>', lambda e: self.edit_transaction())
        
        # 狀態列
        status_frame = ttk.Frame(list_frame)
        status_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(5, 0))
        
        self.status_label = ttk.Label(status_frame, text="準備就緒")
        self.status_label.pack(side=tk.LEFT)
        
        self.record_count_label = ttk.Label(status_frame, text="")
        self.record_count_label.pack(side=tk.RIGHT)
    
    def setup_statistics(self, parent):
        """設定統計區域"""
        stats_frame = ttk.LabelFrame(parent, text="統計資訊", padding="10")
        stats_frame.grid(row=4, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        
        # 統計數據顯示
        stats_data_frame = ttk.Frame(stats_frame)
        stats_data_frame.pack(fill=tk.X)
        
        self.income_label = ttk.Label(stats_data_frame, text="本月收入: $0.00", 
                                     font=("Arial", 11), foreground="green")
        self.income_label.pack(side=tk.LEFT, padx=20)
        
        self.expense_label = ttk.Label(stats_data_frame, text="本月支出: $0.00", 
                                      font=("Arial", 11), foreground="red")
        self.expense_label.pack(side=tk.LEFT, padx=20)
        
        self.balance_label = ttk.Label(stats_data_frame, text="本月結餘: $0.00", 
                                      font=("Arial", 11, "bold"))
        self.balance_label.pack(side=tk.LEFT, padx=20)
        
        # 統計按鈕
        stats_button_frame = ttk.Frame(stats_frame)
        stats_button_frame.pack(side=tk.RIGHT)
        
        ttk.Button(stats_button_frame, text="詳細統計", 
                  command=self.show_detailed_stats).pack(side=tk.LEFT, padx=5)
    
    def setup_reports(self, parent):
        """設定報表區域"""
        report_frame = ttk.LabelFrame(parent, text="統計報表", padding="10")
        report_frame.grid(row=5, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=10)
        
        # 報表類型選擇
        button_frame = ttk.Frame(report_frame)
        button_frame.grid(row=0, column=0, columnspan=4, sticky=(tk.W, tk.E), pady=(0, 10))
        
        self.current_report_type = tk.StringVar(value="year_category")
        
        ttk.Radiobutton(button_frame, text="年分類", variable=self.current_report_type, 
                       value="year_category", command=self.update_report).pack(side=tk.LEFT, padx=10)
        ttk.Radiobutton(button_frame, text="月分類", variable=self.current_report_type, 
                       value="month_category", command=self.update_report).pack(side=tk.LEFT, padx=10)
        ttk.Radiobutton(button_frame, text="月收支", variable=self.current_report_type, 
                       value="month_income_expense", command=self.update_report).pack(side=tk.LEFT, padx=10)
        ttk.Radiobutton(button_frame, text="日收支", variable=self.current_report_type, 
                       value="daily_income_expense", command=self.update_report).pack(side=tk.LEFT, padx=10)
        
        # 時間控制
        control_frame = ttk.Frame(report_frame)
        control_frame.grid(row=1, column=0, columnspan=4, sticky=(tk.W, tk.E), pady=(0, 10))
        
        ttk.Label(control_frame, text="年份:").pack(side=tk.LEFT, padx=(0, 5))
        self.year_var = tk.StringVar(value=str(datetime.now().year))
        year_combo = ttk.Combobox(control_frame, textvariable=self.year_var, width=8, state="readonly")
        
        current_year = datetime.now().year
        year_options = [str(year) for year in range(current_year - 5, current_year + 2)]
        year_combo['values'] = year_options
        year_combo.pack(side=tk.LEFT, padx=(0, 15))
        year_combo.bind('<<ComboboxSelected>>', lambda e: self.update_report())
        
        ttk.Label(control_frame, text="月份:").pack(side=tk.LEFT, padx=(0, 5))
        self.month_var = tk.StringVar(value=str(datetime.now().month))
        month_combo = ttk.Combobox(control_frame, textvariable=self.month_var, width=8, state="readonly")
        month_combo['values'] = [str(i) for i in range(1, 13)]
        month_combo.pack(side=tk.LEFT, padx=(0, 15))
        month_combo.bind('<<ComboboxSelected>>', lambda e: self.update_report())
        
        ttk.Button(control_frame, text="更新報表", command=self.update_report).pack(side=tk.LEFT, padx=15)
        
        # 報表顯示區域
        self.report_display_frame = ttk.Frame(report_frame)
        self.report_display_frame.grid(row=2, column=0, columnspan=4, sticky=(tk.W, tk.E, tk.N, tk.S), pady=10)
        
        report_frame.columnconfigure(3, weight=1)
        report_frame.rowconfigure(2, weight=1)
        
        self.update_report()
    
    def sort_by_column(self, col):
        """依欄位排序"""
        # 這裡可以實作排序功能
        self.status_label.config(text=f"依 {col} 排序")
    
    def filter_current_month(self):
        """篩選本月資料"""
        now = datetime.now()
        start_date = f"{now.year}-{now.month:02d}-01"
        last_day = calendar.monthrange(now.year, now.month)[1]
        end_date = f"{now.year}-{now.month:02d}-{last_day}"
        
        self.start_date_var.set(start_date)
        self.end_date_var.set(end_date)
        self.apply_filters()
    
    def filter_current_week(self):
        """篩選本週資料"""
        from datetime import timedelta
        now = datetime.now()
        week_start = now - timedelta(days=now.weekday())
        week_end = week_start + timedelta(days=6)
        
        self.start_date_var.set(week_start.strftime('%Y-%m-%d'))
        self.end_date_var.set(week_end.strftime('%Y-%m-%d'))
        self.apply_filters()
    
    def filter_today(self):
        """篩選今日資料"""
        today = datetime.now().strftime('%Y-%m-%d')
        self.start_date_var.set(today)
        self.end_date_var.set(today)
        self.apply_filters()
    
    def toggle_filter(self):
        """切換篩選區域顯示/隱藏"""
        if self.filter_expanded.get():
            self.filter_frame.grid_remove()
            self.toggle_button.config(text="▼ 進階篩選")
            self.filter_expanded.set(False)
        else:
            self.filter_frame.grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(5, 10))
            self.toggle_button.config(text="▲ 進階篩選")
            self.filter_expanded.set(True)
    
    def update_category_filter_options(self):
        """更新分類篩選選項"""
        categories = self.category_manager.get_all_categories()
        category_options = ["全部分類"] + [cat['name'] for cat in categories]
        self.category_filter_combo['values'] = category_options
        self.category_filter_combo.set("全部分類")
    
    def clear_filters(self):
        """清除所有篩選條件"""
        self.start_date_var.set("")
        self.end_date_var.set("")
        self.type_filter_var.set("all")
        self.category_filter_var.set("全部分類")
        self.keyword_var.set("")
        self.refresh_data()
    
    def on_filter_change(self, *args):
        """當篩選條件改變時（即時篩選）"""
        self.apply_filters()
    
    def apply_filters(self):
        """套用篩選條件"""
        start_date = self.start_date_var.get().strip()
        end_date = self.end_date_var.get().strip()
        transaction_type = self.type_filter_var.get()
        category_filter = self.category_filter_var.get()
        keyword = self.keyword_var.get().strip().lower()
        
        transactions = self.transaction_manager.get_transactions(limit=1000)
        
        filtered_transactions = []
        for trans in transactions:
            if start_date and trans['date'] < start_date:
                continue
            if end_date and trans['date'] > end_date:
                continue
            if transaction_type != "all" and trans['type'] != transaction_type:
                continue
            if category_filter != "全部分類" and trans['category_name'] != category_filter:
                continue
            if keyword and keyword not in str(trans.get('description', '')).lower():
                continue
            
            filtered_transactions.append(trans)
        
        self.display_filtered_transactions(filtered_transactions)
    
    def display_filtered_transactions(self, transactions):
        """顯示篩選後的交易記錄"""
        for item in self.transaction_tree.get_children():
            self.transaction_tree.delete(item)
        
        self.current_transactions = transactions
        
        for trans in transactions:
            amount_display = f"${trans['amount']:.2f}"
            if trans['type'] == 'income':
                amount_display = f"+{amount_display}"
            else:
                amount_display = f"-{amount_display}"
            
            type_display = "收入" if trans['type'] == 'income' else "支出"
            
            # 添加顏色標記
            tags = ('income' if trans['type'] == 'income' else 'expense',)
            
            self.transaction_tree.insert('', 'end', values=(
                trans['date'],
                type_display,
                trans['category_name'],
                amount_display,
                trans.get('description', '')
            ), tags=(str(trans['id']),) + tags)
        
        # 設定顏色
        self.transaction_tree.tag_configure('income', foreground='green')
        self.transaction_tree.tag_configure('expense', foreground='red')
        
        self.update_filtered_statistics(transactions)
        self.record_count_label.config(text=f"共 {len(transactions)} 筆記錄")
    
    def refresh_data(self):
        """重新整理資料顯示"""
        try:
            for item in self.transaction_tree.get_children():
                self.transaction_tree.delete(item)
            
            transactions = self.transaction_manager.get_transactions(limit=200)
            self.current_transactions = transactions
            
            for trans in transactions:
                amount_display = f"${trans['amount']:.2f}"
                if trans['type'] == 'income':
                    amount_display = f"+{amount_display}"
                else:
                    amount_display = f"-{amount_display}"
                
                type_display = "收入" if trans['type'] == 'income' else "支出"
                tags = ('income' if trans['type'] == 'income' else 'expense',)
                
                self.transaction_tree.insert('', 'end', values=(
                    trans['date'],
                    type_display,
                    trans['category_name'],
                    amount_display,
                    trans.get('description', '')
                ), tags=(str(trans['id']),) + tags)
            
            # 設定顏色
            self.transaction_tree.tag_configure('income', foreground='green')
            self.transaction_tree.tag_configure('expense', foreground='red')
            
            self.update_statistics()
            self.record_count_label.config(text=f"共 {len(transactions)} 筆記錄")
            
            if hasattr(self, 'category_filter_combo'):
                self.update_category_filter_options()
            
            self.update_report()
            self.status_label.config(text="資料已更新")
            
        except Exception as e:
            messagebox.showerror("錯誤", f"資料更新失敗：{e}")
            self.status_label.config(text="更新失敗")
    
    def update_statistics(self):
        """更新統計顯示"""
        now = datetime.now()
        summary = self.transaction_manager.get_monthly_summary(now.year, now.month)
        
        self.income_label.config(text=f"本月收入: ${summary['total_income']:,.2f}")
        self.expense_label.config(text=f"本月支出: ${summary['total_expense']:,.2f}")
        
        balance = summary['balance']
        balance_text = f"本月結餘: ${balance:,.2f}"
        balance_color = "green" if balance >= 0 else "red"
        self.balance_label.config(text=balance_text, foreground=balance_color)
    
    def update_filtered_statistics(self, transactions):
        """更新篩選後的統計顯示"""
        total_income = sum(trans['amount'] for trans in transactions if trans['type'] == 'income')
        total_expense = sum(trans['amount'] for trans in transactions if trans['type'] == 'expense')
        balance = total_income - total_expense
        
        self.income_label.config(text=f"篩選收入: ${total_income:,.2f}")
        self.expense_label.config(text=f"篩選支出: ${total_expense:,.2f}")
        
        balance_text = f"篩選結餘: ${balance:,.2f}"
        balance_color = "green" if balance >= 0 else "red"
        self.balance_label.config(text=balance_text, foreground=balance_color)
    
    def show_detailed_stats(self):
        """顯示詳細統計"""
        stats_window = tk.Toplevel(self.root)
        stats_window.title("詳細統計")
        stats_window.geometry("600x400")
        stats_window.transient(self.root)
        
        # 這裡可以添加更詳細的統計資訊
        ttk.Label(stats_window, text="詳細統計功能開發中...", 
                 font=("Arial", 12)).pack(expand=True)
    
    def update_report(self):
        """更新報表顯示"""
        report_type = self.current_report_type.get()
        
        # 清除現有報表
        for widget in self.report_display_frame.winfo_children():
            widget.destroy()
        
        if not MATPLOTLIB_AVAILABLE:
            ttk.Label(self.report_display_frame, text="圖表功能需要安裝 matplotlib", 
                     font=("Arial", 12), foreground="red").pack(pady=20)
            ttk.Label(self.report_display_frame, text="執行: pip install matplotlib").pack()
            return
        
        try:
            if report_type == "year_category":
                self.show_year_category_chart()
            elif report_type == "month_category":
                self.show_month_category_chart()
            elif report_type == "month_income_expense":
                self.show_month_income_expense_chart()
            elif report_type == "daily_income_expense":
                self.show_daily_income_expense_chart()
        except Exception as e:
            ttk.Label(self.report_display_frame, text=f"報表生成失敗: {e}", 
                     foreground="red").pack(pady=20)
    
    def show_year_category_chart(self):
        """顯示年度分類圓餅圖"""
        year = int(self.year_var.get())
        start_date = f"{year}-01-01"
        end_date = f"{year}-12-31"
        transactions = self.transaction_manager.get_transactions_by_date_range(start_date, end_date)

        if not transactions:
            ttk.Label(self.report_display_frame, text="本年無交易資料").pack(pady=20)
            return

        # 統計分類資料
        expense_stats = defaultdict(float)
        income_stats = defaultdict(float)

        for trans in transactions:
            if trans['type'] == 'expense':
                expense_stats[trans['category_name']] += trans['amount']
            else:
                income_stats[trans['category_name']] += trans['amount']

        # 建立圖表
        fig = Figure(figsize=(10, 5), dpi=80)
        ax = fig.add_subplot(111)

        if expense_stats:
            sorted_expenses = sorted(expense_stats.items(), key=lambda x: x[1], reverse=True)
            labels = [cat for cat, _ in sorted_expenses]
            sizes = [amount for _, amount in sorted_expenses]

            colors = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FECA57', '#FF9FF3', '#54A0FF', '#5F27CD']
            ax.pie(sizes, labels=labels, colors=colors[:len(labels)], autopct='%1.1f%%', startangle=90)
            ax.set_title(f'{year}年支出分類', fontsize=12, fontweight='bold')
        else:
            ttk.Label(self.report_display_frame, text="本年無支出資料").pack(pady=20)
            return

        fig.tight_layout()

        canvas = FigureCanvasTkAgg(fig, self.report_display_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
    
    # 交易管理方法
    def add_transaction(self):
        """新增交易記錄"""
        dialog = TransactionDialog(self.root, self.category_manager, self.transaction_manager)
        self.root.wait_window(dialog.dialog)
        
        if dialog.result:
            data = dialog.result
            success = self.transaction_manager.add_transaction(
                date=data['date'],
                transaction_type=data['type'],
                category_id=data['category_id'],
                amount=data['amount'],
                description=data['description']
            )
            
            if success:
                messagebox.showinfo("成功", "交易記錄新增成功！")
                self.refresh_data()
                self.status_label.config(text="新增記錄成功")
            else:
                messagebox.showerror("錯誤", "交易記錄新增失敗！")
                self.status_label.config(text="新增記錄失敗")
    
    def edit_transaction(self):
        """編輯選中的交易記錄"""
        selected_item = self.transaction_tree.selection()
        if not selected_item:
            messagebox.showwarning("提醒", "請先選擇要編輯的交易記錄")
            return
        
        transaction_id = int(self.transaction_tree.item(selected_item[0])['tags'][0])
        
        transactions = self.transaction_manager.get_transactions()
        transaction_data = None
        for trans in transactions:
            if trans['id'] == transaction_id:
                transaction_data = trans
                break
        
        if not transaction_data:
            messagebox.showerror("錯誤", "找不到交易記錄")
            return
        
        dialog = TransactionDialog(self.root, self.category_manager, 
                                 self.transaction_manager, transaction_data)
        self.root.wait_window(dialog.dialog)
        
        if dialog.result:
            data = dialog.result
            success = self.transaction_manager.update_transaction(
                transaction_id=transaction_id,
                date=data['date'],
                transaction_type=data['type'],
                category_id=data['category_id'],
                amount=data['amount'],
                description=data['description']
            )
            
            if success:
                messagebox.showinfo("成功", "交易記錄更新成功！")
                self.refresh_data()
                self.status_label.config(text="更新記錄成功")
            else:
                messagebox.showerror("錯誤", "交易記錄更新失敗！")
                self.status_label.config(text="更新記錄失敗")
    
    def delete_transaction(self):
        """刪除選中的交易記錄"""
        selected_item = self.transaction_tree.selection()
        if not selected_item:
            messagebox.showwarning("提醒", "請先選擇要刪除的交易記錄")
            return
        
        if not messagebox.askyesno("確認", "確定要刪除這筆交易記錄嗎？\n此操作無法復原。"):
            return
        
        transaction_id = int(self.transaction_tree.item(selected_item[0])['tags'][0])
        
        success = self.transaction_manager.delete_transaction(transaction_id)
        
        if success:
            messagebox.showinfo("成功", "交易記錄刪除成功！")
            self.refresh_data()
            self.status_label.config(text="刪除記錄成功")
        else:
            messagebox.showerror("錯誤", "交易記錄刪除失敗！")
            self.status_label.config(text="刪除記錄失敗")
    
    # 匯出功能方法
    def export_to_csv(self):
        """匯出資料到 CSV 檔案"""
        if not self.current_transactions:
            messagebox.showwarning("提醒", "沒有資料可匯出")
            return
        
        filename = filedialog.asksaveasfilename(
            title="匯出 CSV 檔案",
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            initialname=f"記帳資料_{datetime.now().strftime('%Y%m%d')}.csv"
        )
        
        if not filename:
            return
        
        try:
            with open(filename, 'w', newline='', encoding='utf-8-sig') as csvfile:
                writer = csv.writer(csvfile)
                
                # 寫入標題
                writer.writerow(['日期', '類型', '分類', '金額', '備註'])
                
                # 寫入交易資料
                for trans in self.current_transactions:
                    type_display = "收入" if trans['type'] == 'income' else "支出"
                    writer.writerow([
                        trans['date'],
                        type_display,
                        trans['category_name'],
                        trans['amount'],
                        trans.get('description', '')
                    ])
                
                # 寫入統計摘要
                writer.writerow([])
                writer.writerow(['統計摘要'])
                
                total_income = sum(trans['amount'] for trans in self.current_transactions if trans['type'] == 'income')
                total_expense = sum(trans['amount'] for trans in self.current_transactions if trans['type'] == 'expense')
                balance = total_income - total_expense
                
                writer.writerow(['總收入', f'${total_income:.2f}'])
                writer.writerow(['總支出', f'${total_expense:.2f}'])
                writer.writerow(['結餘', f'${balance:.2f}'])
                
                # 寫入匯出資訊
                writer.writerow([])
                writer.writerow(['匯出資訊'])
                writer.writerow(['匯出時間', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
                writer.writerow(['記錄筆數', len(self.current_transactions)])
            
            messagebox.showinfo("成功", f"資料已成功匯出到：\n{filename}")
            self.status_label.config(text="CSV 匯出成功")
            
        except Exception as e:
            messagebox.showerror("錯誤", f"匯出失敗：{str(e)}")
            self.status_label.config(text="CSV 匯出失敗")
    
    def export_to_excel(self):
        """匯出資料到 Excel 檔案"""
        if not self.current_transactions:
            messagebox.showwarning("提醒", "沒有資料可匯出")
            return
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            result = messagebox.askyesno("缺少套件", 
                "Excel 匯出需要安裝 openpyxl 套件。\n\n" +
                "請在終端機執行：pip install openpyxl\n\n" +
                "現在要改用 CSV 格式匯出嗎？")
            if result:
                self.export_to_csv()
            return
        
        filename = filedialog.asksaveasfilename(
            title="匯出 Excel 檔案",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialname=f"記帳資料_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
        
        if not filename:
            return
        
        try:
            wb = openpyxl.Workbook()
            ws_data = wb.active
            ws_data.title = "交易記錄"
            
            # 設定標題樣式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            headers = ['日期', '類型', '分類', '金額', '備註']
            for col, header in enumerate(headers, 1):
                cell = ws_data.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # 寫入交易資料
            for row, trans in enumerate(self.current_transactions, 2):
                ws_data.cell(row=row, column=1, value=trans['date'])
                ws_data.cell(row=row, column=2, value="收入" if trans['type'] == 'income' else "支出")
                ws_data.cell(row=row, column=3, value=trans['category_name'])
                ws_data.cell(row=row, column=4, value=trans['amount'])
                ws_data.cell(row=row, column=5, value=trans.get('description', ''))
            
            # 調整欄寬
            column_widths = [12, 8, 15, 12, 30]
            for col, width in enumerate(column_widths, 1):
                ws_data.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
            
            wb.save(filename)
            messagebox.showinfo("成功", f"Excel 檔案已成功匯出到：\n{filename}")
            self.status_label.config(text="Excel 匯出成功")
            
        except Exception as e:
            messagebox.showerror("錯誤", f"Excel 匯出失敗：{str(e)}")
            self.status_label.config(text="Excel 匯出失敗")
    
    def on_closing(self):
        """程式關閉時的處理"""
        if messagebox.askokcancel("退出", "確定要退出個人記帳本嗎？"):
            self.root.destroy()
    
    def run(self):
        """啟動主程式"""
        # 綁定關閉事件
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        # 顯示啟動訊息
        self.status_label.config(text="個人記帳本已啟動")
        
        print("🚀 個人記帳本已啟動")
        print("📚 使用說明：")
        print("   - Ctrl+N: 新增交易")
        print("   - Ctrl+E: 編輯交易")
        print("   - Del: 刪除交易")
        print("   - F5: 重新整理")
        print("   - Ctrl+S: 匯出 CSV")
        
        # 啟動主迴圈
        self.root.mainloop()

def main():
    """主程式入口"""
    try:
        print("正在啟動個人記帳本...")
        app = MainWindow()
        app.run()
    except KeyboardInterrupt:
        print("\n程式被使用者中斷")
    except Exception as e:
        print(f"程式執行錯誤：{e}")
        import traceback
        traceback.print_exc()
        input("按 Enter 鍵退出...")

        transactions = self.transaction_manager.get_transactions_by_date_range(start_date, end_date)
        
        if not transactions:
            ttk.Label(self.report_display_frame, text="無資料可顯示").pack(pady=20)
            return
        
        # 統計資料
        expense_stats = defaultdict(float)
        income_stats = defaultdict(float)
        
        for trans in transactions:
            if trans['type'] == 'expense':
                expense_stats[trans['category_name']] += trans['amount']
            else:
                income_stats[trans['category_name']] += trans['amount']
        
        # 建立圖表
        fig = Figure(figsize=(12, 6), dpi=80)
        
        if expense_stats:
            ax1 = fig.add_subplot(121)
            
            sorted_expenses = sorted(expense_stats.items(), key=lambda x: x[1], reverse=True)
            if len(sorted_expenses) > 8:
                main_expenses = sorted_expenses[:7]
                other_amount = sum(amount for _, amount in sorted_expenses[7:])
                main_expenses.append(('其他', other_amount))
            else:
                main_expenses = sorted_expenses
            
            labels = [cat for cat, _ in main_expenses]
            sizes = [amount for _, amount in main_expenses]
            colors = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FECA57', '#FF9FF3', '#54A0FF', '#5F27CD']
            
            ax1.pie(sizes, labels=labels, colors=colors[:len(labels)], autopct='%1.1f%%', startangle=90)
            ax1.set_title(f'{year}年支出分類', fontsize=12, fontweight='bold')
        
        if income_stats:
            ax2 = fig.add_subplot(122)
            
            sorted_income = sorted(income_stats.items(), key=lambda x: x[1], reverse=True)
            labels = [cat for cat, _ in sorted_income]
            sizes = [amount for _, amount in sorted_income]
            colors = ['#28a745', '#20c997', '#17a2b8', '#6f42c1', '#e83e8c', '#fd7e14']
            
            ax2.pie(sizes, labels=labels, colors=colors[:len(labels)], autopct='%1.1f%%', startangle=90)
            ax2.set_title(f'{year}年收入分類', fontsize=12, fontweight='bold')
        
        fig.tight_layout()
        
        canvas = FigureCanvasTkAgg(fig, self.report_display_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        
        # 統計摘要
        total_income = sum(income_stats.values())
        total_expense = sum(expense_stats.values())
        
        summary_frame = ttk.Frame(self.report_display_frame)
        summary_frame.pack(fill=tk.X, pady=(10, 0))
        
        ttk.Label(summary_frame, text=f"總收入: ${total_income:,.2f}", 
                 foreground="green").pack(side=tk.LEFT, padx=20)
        ttk.Label(summary_frame, text=f"總支出: ${total_expense:,.2f}", 
                 foreground="red").pack(side=tk.LEFT, padx=20)
        ttk.Label(summary_frame, text=f"結餘: ${total_income - total_expense:,.2f}").pack(side=tk.LEFT, padx=20)
    
    def show_month_category_chart(self):
        """顯示月度分類圓餅圖"""
        year = int(self.year_var.get())
        month = int(self.month_var.get())
        
        start_date = f"{year}-{month:02d}-01"
        last_day = calendar.monthrange(year, month)[1]
        end_date = f"{year}-{month:02d}-{last_day}"
        
        transactions = self.transaction_manager.get_transactions_by_date_range(start_date, end_date)
        
        if not transactions:
            ttk.Label(self.report_display_frame, text="無資料可顯示").pack(pady=20)
            return
        
        # 統計資料
        expense_stats = defaultdict(float)
        income_stats = defaultdict(float)
        
        for trans in transactions:
            if trans['type'] == 'expense':
                expense_stats[trans['category_name']] += trans['amount']
            else:
                income_stats[trans['category_name']] += trans['amount']
        
        # 建立圖表（類似年度圖表的邏輯）
        fig = Figure(figsize=(10, 5), dpi=80)
        ax = fig.add_subplot(111)
        
        if expense_stats:
            sorted_expenses = sorted(expense_stats.items(), key=lambda x: x[1], reverse=True)
            labels = [cat for cat, _ in sorted_expenses]
            sizes = [amount for _, amount in sorted_expenses]
            
            colors = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FECA57', '#FF9FF3', '#54A0FF', '#5F27CD']
            ax.pie(sizes, labels=labels, colors=colors[:len(labels)], autopct='%1.1f%%', startangle=90)
            ax.set_title(f'{year}年{month}月支出分類', fontsize=12, fontweight='bold')
        
        fig.tight_layout()
        
        canvas = FigureCanvasTkAgg(fig, self.report_display_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
    
    def show_month_income_expense_chart(self):
        """顯示月度收支長條圖"""
        year = int(self.year_var.get())
        
        # 收集12個月的資料
        months_labels = []
        income_data = []
        expense_data = []
        
        for month in range(1, 13):
            summary = self.transaction_manager.get_monthly_summary(year, month)
            
            if summary['total_income'] > 0 or summary['total_expense'] > 0:
                months_labels.append(f"{month}月")
                income_data.append(summary['total_income'])
                expense_data.append(summary['total_expense'])
        
        if not months_labels:
            ttk.Label(self.report_display_frame, text="無資料可顯示").pack(pady=20)
            return
        
        # 建立圖表
        fig = Figure(figsize=(12, 6), dpi=80)
        ax = fig.add_subplot(111)
        
        x = range(len(months_labels))
        width = 0.35
        
        bars1 = ax.bar([i - width/2 for i in x], income_data, width, label='收入', color='#28a745', alpha=0.8)
        bars2 = ax.bar([i + width/2 for i in x], expense_data, width, label='支出', color='#dc3545', alpha=0.8)
        
        # 在柱子上顯示數值
        for bar in bars1:
            height = bar.get_height()
            if height > 0:
                ax.text(bar.get_x() + bar.get_width()/2., height + max(income_data + expense_data) * 0.01,
                       f'${height:.0f}', ha='center', va='bottom', fontsize=8)
        
        for bar in bars2:
            height = bar.get_height()
            if height > 0:
                ax.text(bar.get_x() + bar.get_width()/2., height + max(income_data + expense_data) * 0.01,
                       f'${height:.0f}', ha='center', va='bottom', fontsize=8)
        
        ax.set_title(f'{year}年月度收支對比', fontsize=14, fontweight='bold')
        ax.set_xlabel('月份', fontsize=12)
        ax.set_ylabel('金額 (元)', fontsize=12)
        ax.set_xticks(x)
        ax.set_xticklabels(months_labels)
        ax.legend()
        ax.grid(True, alpha=0.3, axis='y')
        
        fig.tight_layout()
        
        canvas = FigureCanvasTkAgg(fig, self.report_display_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
    
    def show_daily_income_expense_chart(self):
        """顯示日度收支長條圖"""
        year = int(self.year_var.get())
        month = int(self.month_var.get())
        
        start_date = f"{year}-{month:02d}-01"
        last_day = calendar.monthrange(year, month)[1]
        end_date = f"{year}-{month:02d}-{last_day}"
        
        transactions = self.transaction_manager.get_transactions_