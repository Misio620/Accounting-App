"""
主視窗模組 - 個人記帳應用程式主界面
重構版本 - 使用模組化設計
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import customtkinter as ctk
from datetime import datetime
import sys
import os
import csv

# 匯入資料庫模組
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from database.models import DatabaseManager, CategoryManager, TransactionManager

# 匯入 GUI 模組
from .dialogs import TransactionDialog, CategoryManagementDialog
from .charts import ChartManager, MATPLOTLIB_AVAILABLE
from .filters import FilterPanel
from .ui_config import COLORS, FONTS, SPACING, PADDING, ICONS
from .ui_components import StatCard, ModernButton, SectionFrame

# 匯入工具模組
try:
    sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
    from utils.backup import BackupManager, format_file_size
    BACKUP_AVAILABLE = True
except ImportError:
    BACKUP_AVAILABLE = False


class MainWindow:
    """主視窗類別 - 重構版本 (CustomTkinter)"""
    
    def __init__(self):
        print("正在初始化個人記帳本...")
        
        # 初始化資料庫
        try:
            self.db_manager = DatabaseManager("accounting.db")
            self.category_manager = CategoryManager(self.db_manager)
            self.transaction_manager = TransactionManager(self.db_manager)
            print("✅ 資料庫初始化完成")
        except Exception as e:
            print(f"❌ 資料庫初始化失敗: {e}")
            messagebox.showerror("資料庫錯誤", f"無法初始化資料庫：{e}")
            sys.exit(1)
        
        # 初始化圖表管理器
        self.chart_manager = ChartManager(self.transaction_manager)
        
        # 初始化備份管理器
        if BACKUP_AVAILABLE:
            self.backup_manager = BackupManager()
        else:
            self.backup_manager = None
        
        # 建立主視窗 (CustomTkinter)
        ctk.set_appearance_mode("Light")  # 極簡白風格
        ctk.set_default_color_theme("blue")
        
        self.root = ctk.CTk()
        self.root.title("個人記帳本 v2.0 (Modern UI)")
        self.root.geometry("1100x850") # 稍微加大以適應寬鬆排版
        self.root.minsize(900, 700)
        
        # 設定全域字體大致比例 (CTk 會自動縮放，但這裡保留參考)
        # self.root.option_add("*Font", FONTS['body']) # CTk 不吃這個，但 tk 元件 (如 Treeview) 吃
        
        self.current_transactions = []
        
        self.setup_ui()
        
        print("✅ 界面初始化完成")
    
    def setup_ui(self):
        """設定主界面 (Dashboard Layout)"""
        # self.setup_menu() - 已移除，改用 Sidebar + Settings View
        
        # 主佈局配置
        self.root.grid_columnconfigure(0, weight=0) # Sidebar 固定
        self.root.grid_columnconfigure(1, weight=1) # Content 自適應
        self.root.grid_rowconfigure(0, weight=1)
        
        # 視圖管理初始化
        self.views = {}
        self.nav_buttons = {}
        
        # 1. 建立側邊欄 Sidebar
        self.setup_sidebar()
        
        # 2. 建立內容區域 Content Area
        self.content_area = ctk.CTkFrame(self.root, fg_color="transparent", corner_radius=0)
        self.content_area.grid(row=0, column=1, sticky="nsew", padx=SPACING['lg'], pady=SPACING['lg'])
        self.content_area.grid_rowconfigure(0, weight=1)
        self.content_area.grid_columnconfigure(0, weight=1)
        
        # 綁定快捷鍵
        self.setup_shortcuts()
        
        # 初始顯示 Dashboard
        self.root.after(100, lambda: self.switch_view('dashboard'))

    # setup_menu 已移除

    def setup_sidebar(self):
        """建立左側導航欄"""
        self.sidebar = ctk.CTkFrame(self.root, fg_color=COLORS['sidebar_bg'], width=240, corner_radius=0)
        self.sidebar.grid(row=0, column=0, sticky="nsew")
        self.sidebar.grid_rowconfigure(4, weight=1) # Spacer push bottom
        
        # Logo / App Name
        logo_label = ctk.CTkLabel(
            self.sidebar,
            text=f" {ICONS['balance']} 個人記帳本",
            font=(FONTS['heading'][0], 20, "bold"),
            text_color="#FFFFFF"
        )
        logo_label.grid(row=0, column=0, padx=20, pady=(20, 10))
        
        version_label = ctk.CTkLabel(
            self.sidebar,
            text="v2.0 Dashboard",
            font=(FONTS['caption'][0], 12),
            text_color=COLORS['text_secondary']
        )
        version_label.grid(row=1, column=0, padx=20, pady=(0, 20))
        
        # CTA Button (記一筆)
        cta_btn = ModernButton(
            self.sidebar,
            text="記一筆",
            icon='add',
            style='primary',
            height=40,
            command=self.add_transaction
        )
        cta_btn.grid(row=2, column=0, padx=20, pady=10, sticky="ew")
        
        # Navigation
        self.nav_frame = ctk.CTkFrame(self.sidebar, fg_color="transparent")
        self.nav_frame.grid(row=3, column=0, sticky="ew", pady=10)
        
        # 主導航
        self.create_nav_button("dashboard", f"{ICONS['chart']} 首頁", self.nav_frame)
        
        # 分隔線 (字體與 nav button 統一)
        ctk.CTkLabel(self.nav_frame, text="── 報表分析 ──", text_color=COLORS['text_secondary'], 
                     font=(FONTS['body'][0], 13, "bold")).pack(fill="x", padx=15, pady=(15, 5))
        
        # 報表快捷按鈕
        self.create_report_button("year_category", "📊 年分類", self.nav_frame)
        self.create_report_button("month_category", "📊 月分類", self.nav_frame)
        self.create_report_button("month_income_expense", "📈 月收支", self.nav_frame)
        self.create_report_button("daily_income_expense", "📈 日收支", self.nav_frame)
        
        # 分隔線
        ctk.CTkLabel(self.nav_frame, text="", text_color=COLORS['text_secondary']).pack(fill="x", pady=5)
        
        # 資料管理
        self.create_nav_button("settings", f"{ICONS['settings']} 資料管理", self.nav_frame)
        
        # Bottom Area
        self.sidebar_bottom = ctk.CTkFrame(self.sidebar, fg_color="transparent")
        self.sidebar_bottom.grid(row=5, column=0, sticky="ew", pady=20)
        
        self.status_label = ctk.CTkLabel(
            self.sidebar_bottom, 
            text="準備就緒", 
            text_color=COLORS['text_secondary'],
            font=(FONTS['caption'][0], 10)
        )
        self.status_label.pack(side="bottom", pady=5)
        
    def create_nav_button(self, view_name, text, parent):
        """建立導航按鈕"""
        btn = ctk.CTkButton(
            parent,
            text=text,
            height=40,
            corner_radius=5,
            border_spacing=10,
            text_color=COLORS['sidebar_text'],
            fg_color="transparent",
            hover_color=COLORS['sidebar_hover'],
            anchor="w",
            font=(FONTS['body'][0], 13, "bold"),
            command=lambda: self.switch_view(view_name)
        )
        btn.pack(fill="x", padx=10, pady=2)
        self.nav_buttons[view_name] = btn
    
    def create_report_button(self, report_type, text, parent):
        """建立報表快捷按鈕"""
        view_name = f"report_{report_type}"
        btn = ctk.CTkButton(
            parent,
            text=text,
            height=40,
            corner_radius=5,
            border_spacing=10,
            text_color=COLORS['sidebar_text'],
            fg_color="transparent",
            hover_color=COLORS['sidebar_hover'],
            anchor="w",
            font=(FONTS['body'][0], 13, "bold"),
            command=lambda: self.switch_view(view_name)
        )
        btn.pack(fill="x", padx=10, pady=2)
        self.nav_buttons[view_name] = btn
        
    def switch_view(self, view_name):
        """切換視圖"""
        # 更新按鈕狀態
        for name, btn in self.nav_buttons.items():
            if name == view_name:
                btn.configure(
                    fg_color=COLORS['sidebar_selected'],
                    text_color=COLORS['sidebar_text_active']
                )
            else:
                btn.configure(
                    fg_color="transparent",
                    text_color=COLORS['sidebar_text']
                )
        
        # 隱藏當前視圖
        if hasattr(self, 'current_view_frame') and self.current_view_frame:
            self.current_view_frame.pack_forget()
            
        # 顯示/建立目標視圖
        if view_name not in self.views:
            # Lazy loading
            self.create_view(view_name)
            
        self.current_view_frame = self.views[view_name]
        self.current_view_frame.pack(fill="both", expand=True)
        
        # 觸發特定視圖的刷新邏輯
        if view_name == 'dashboard':
            self.refresh_dashboard()
        elif view_name == 'transactions':
            self.refresh_transactions()
        elif view_name.startswith('report_'):
            # 報表視圖：更新當前報表類型並刷新
            report_type = view_name.replace('report_', '')
            self.current_report_type = report_type
            self.current_report_parent = self.views[view_name]
            # 延遲刷新以確保視圖已顯示
            self.root.after(50, self._refresh_current_chart)
            
    def create_view(self, view_name):
        """工廠方法：建立各個視圖"""
        # 每個 View 都是一個 CTkFrame，背景預設為 bg_primary (content color)
        view = ctk.CTkFrame(self.content_area, fg_color="transparent")
        self.views[view_name] = view
        
        if view_name == 'dashboard':
            self._setup_view_dashboard(view)
        elif view_name == 'transactions':
            self._setup_view_transactions(view)
        elif view_name == 'reports':
            self._setup_view_reports(view)
        elif view_name == 'settings':
            self._setup_view_settings(view)
        elif view_name.startswith('report_'):
            # 內嵌報表視圖
            report_type = view_name.replace('report_', '')
            self._setup_view_report_embed(view, report_type)

    
    # setup_menu - 已移除

    
    def setup_shortcuts(self):
        """設定快捷鍵"""
        self.root.bind('<Control-n>', lambda e: self.add_transaction())
        self.root.bind('<F5>', lambda e: self.refresh_data())
        self.root.bind('<Control-s>', lambda e: self.export_to_csv())
    
    # --- View Setup Methods ---
    
    def _setup_view_dashboard(self, parent):
        """Dashboard View: 統計卡片與交易列表"""
        # Header
        header = ctk.CTkFrame(parent, fg_color="transparent")
        header.pack(fill="x", pady=(0, 15))
        
        ctk.CTkLabel(header, text="首頁", font=(FONTS['title'][0], 24, "bold")).pack(side="left")
        ctk.CTkLabel(header, text=f"{datetime.now().strftime('%Y年%m月%d日')}", 
                   font=(FONTS['body'][0], 14), text_color=COLORS['text_secondary']).pack(side="right", anchor="s")
        
        # 1. 統計卡片區域
        stats_frame = ctk.CTkFrame(parent, fg_color="transparent")
        stats_frame.pack(fill="x", pady=(0, 15))
        stats_frame.grid_columnconfigure(0, weight=1)
        stats_frame.grid_columnconfigure(1, weight=1)
        stats_frame.grid_columnconfigure(2, weight=1)
        
        self.income_card = StatCard(stats_frame, card_type='income')
        self.income_card.grid(row=0, column=0, sticky="ew", padx=(0, 10))
        
        self.expense_card = StatCard(stats_frame, card_type='expense')
        self.expense_card.grid(row=0, column=1, sticky="ew", padx=10)
        
        self.balance_card = StatCard(stats_frame, card_type='balance')
        self.balance_card.grid(row=0, column=2, sticky="ew", padx=(10, 0))
        
        # 2. 快速篩選按鈕列
        filter_bar = ctk.CTkFrame(parent, fg_color="transparent")
        filter_bar.pack(fill="x", pady=(0, 10))
        
        ctk.CTkLabel(filter_bar, text="期間篩選：", font=(FONTS['body'][0], 13)).pack(side="left", padx=(0, 10))
        
        self.period_buttons = {}
        periods = [("today", "本日"), ("week", "本週"), ("month", "本月"), ("year", "本年"), ("all", "所有紀錄")]
        for key, label in periods:
            btn = ctk.CTkButton(
                filter_bar, text=label, width=70, height=32,
                fg_color=COLORS['primary'] if key == "month" else "transparent",
                text_color="white" if key == "month" else COLORS['text_primary'],
                border_width=1, border_color=COLORS['border'],
                corner_radius=5,
                command=lambda k=key: self.filter_by_period(k)
            )
            btn.pack(side="left", padx=3)
            self.period_buttons[key] = btn
        
        self.current_period = "month"  # 預設本月
        
        # 3. 交易列表區域
        list_container = ctk.CTkFrame(parent, fg_color=COLORS['bg_card'], corner_radius=10)
        list_container.pack(fill="both", expand=True)
        
        # Treeview
        self._create_transaction_tree(list_container)

    def _setup_view_transactions(self, parent):
        """Transactions View: 完整交易列表"""
        # Header
        ctk.CTkLabel(parent, text="交易明細", font=(FONTS['title'][0], 24, "bold")).pack(anchor="w", pady=(0, 20))
        
        # 1. 篩選與工具列 (使用 Grid 因為 FilterPanel 設計為 Grid 佈局)
        # 我們創建一個容器來容納 FilterPanel，並加上操作按鈕
        tools_frame = ctk.CTkFrame(parent, fg_color="transparent")
        tools_frame.pack(fill="x", pady=(0, 10))
        tools_frame.grid_columnconfigure(0, weight=1)
        
        # 篩選器
        filter_container = ctk.CTkFrame(tools_frame, fg_color=COLORS['bg_primary'])
        filter_container.grid(row=0, column=0, sticky="ew")
        
        # 按鈕區 (放在篩選器上方或旁? Dashboard 風格通常篩選器常駐)
        # 我們將 FilterPanel 整合進來
        # 注意：FilterPanel 原本期望 control_parent 參數
        
        # 右側功能按鈕 (新增、編輯、匯出...)
        # 這裡我們簡化，只放 Filter。新增按鈕已在 Sidebar。
        
        self.filter_panel = FilterPanel(filter_container, self.category_manager, self.on_filter_applied)
        
        # 2. 交易列表 (Treeview)
        list_container = ctk.CTkFrame(parent, fg_color=COLORS['bg_card'], corner_radius=10)
        list_container.pack(fill="both", expand=True)
        
        # Treeview Style & Setup (Reuse Logic)
        self._create_transaction_tree(list_container)
        
    def _create_transaction_tree(self, parent):
        """建立 Treeview (獨立方法以供複用)"""
        columns = ('date', 'type', 'category', 'amount', 'description')
        
        # Treeview Container
        tree_frame = ctk.CTkFrame(parent, fg_color=COLORS['bg_card'])
        tree_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Scrollbars
        v_scrollbar = ctk.CTkScrollbar(tree_frame, orientation="vertical")
        h_scrollbar = ctk.CTkScrollbar(tree_frame, orientation="horizontal")
        
        # Style
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview", 
                        background="#FFFFFF", 
                        fieldbackground="#FFFFFF", 
                        foreground=COLORS['text_primary'],
                        rowheight=48, # 加高以配合字體
                        borderwidth=0,
                        font=(FONTS['body'][0], 16))
        style.configure("Treeview.Heading", 
                        background="#F8FAFC", 
                        foreground=COLORS['text_secondary'], 
                        relief="flat",
                        font=(FONTS['body'][0], 16, "bold"))
        style.map("Treeview", background=[('selected', COLORS['primary'])], foreground=[('selected', 'white')])

        self.transaction_tree = ttk.Treeview(tree_frame, columns=columns, show='headings', 
                                           yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        v_scrollbar.configure(command=self.transaction_tree.yview)
        h_scrollbar.configure(command=self.transaction_tree.xview)
        
        # Define Columns
        self.transaction_tree.heading('date', text='日期')
        self.transaction_tree.column('date', width=100, anchor='center')
        
        self.transaction_tree.heading('type', text='類型')
        self.transaction_tree.column('type', width=60, anchor='center')
        
        self.transaction_tree.heading('category', text='分類')
        self.transaction_tree.column('category', width=120, anchor='center')
        
        self.transaction_tree.heading('amount', text='金額')
        self.transaction_tree.column('amount', width=100, anchor='center')
        
        self.transaction_tree.heading('description', text='備註')
        self.transaction_tree.column('description', width=300, anchor='center')
        
        # Layout
        self.transaction_tree.pack(side="left", fill="both", expand=True)
        v_scrollbar.pack(side="right", fill="y")
        h_scrollbar.pack(side="bottom", fill="x") # Wait, pack bottom of frame? No. grid is better for scrollbars.
        
        # Re-do layout with grid
        self.transaction_tree.pack_forget()
        v_scrollbar.pack_forget()
        h_scrollbar.pack_forget()
        
        tree_frame.grid_columnconfigure(0, weight=1)
        tree_frame.grid_rowconfigure(0, weight=1)
        
        self.transaction_tree.grid(row=0, column=0, sticky="nsew")
        v_scrollbar.grid(row=0, column=1, sticky="ns")
        h_scrollbar.grid(row=1, column=0, sticky="ew")
        
        # Events
        self.transaction_tree.bind('<Double-1>', lambda e: self.edit_transaction())
        
        # Status / Counts setup logic should be added here or in refresh.
        # Let's add a status bar at bottom of list container
        self.list_status_label = ctk.CTkLabel(parent, text="準備就緒", text_color=COLORS['text_light'], anchor="e")
        self.list_status_label.pack(fill="x", padx=10, pady=(0, 5))

    def _setup_view_reports(self, parent):
        """Reports View: 進階分析入口"""
        ctk.CTkLabel(parent, text="報表分析", font=(FONTS['title'][0], 24, "bold")).pack(anchor="w", pady=(0, 20))
        
        # 網格佈局容器
        grid = ctk.CTkFrame(parent, fg_color="transparent")
        grid.pack(fill="both", expand=True)
        grid.grid_columnconfigure(0, weight=1)
        grid.grid_columnconfigure(1, weight=1)
        
        # 定義按鈕參數 (文字, report_type, row, col, color)
        buttons = [
            ("年度分類分析", "year_category", 0, 0, COLORS['primary']),
            ("月度分類佔比", "month_category", 0, 1, COLORS['info']),
            ("月度收支趨勢", "month_income_expense", 1, 0, COLORS['success']),
            ("每日收支明細", "daily_income_expense", 1, 1, COLORS['warning'])
        ]
        
        for text, report_type, row, col, color in buttons:
            btn_frame = ctk.CTkFrame(grid, fg_color=COLORS['bg_card'], corner_radius=15)
            btn_frame.grid(row=row, column=col, sticky="nsew", padx=10, pady=10)
            
            # 大按鈕
            btn = ctk.CTkButton(
                btn_frame,
                text=text,
                font=(FONTS['heading'][0], 18, "bold"),
                fg_color=color,
                corner_radius=10,
                height=100,
                command=lambda t=report_type: self.open_report_window(t)
            )
            btn.pack(fill="both", expand=True, padx=20, pady=20)
            
            ctk.CTkLabel(btn_frame, text=f"檢視 {text} 報表", text_color=COLORS['text_secondary']).pack(pady=(0, 15))

    def _setup_view_settings(self, parent):
        """Settings / Data View"""
        ctk.CTkLabel(parent, text="資料管理", font=(FONTS['title'][0], 24, "bold")).pack(anchor="w", pady=(0, 20))
        
        # 1. 匯出區
        export_section = SectionFrame(parent, title="資料匯出")
        export_section.pack(fill="x", pady=(0, 20))
        
        ctk.CTkLabel(export_section.content, text="將交易記錄匯出為 CSV 或 Excel 檔案。").pack(anchor="w", pady=(0, 10))
        
        h_box = ctk.CTkFrame(export_section.content, fg_color="transparent")
        h_box.pack(fill="x")
        ModernButton(h_box, text="匯出 CSV", icon='export', command=self.export_to_csv).pack(side="left", padx=(0, 10))
        ModernButton(h_box, text="匯出 Excel", icon='export', style='secondary', command=self.export_to_excel).pack(side="left")
        
        # 2. 備份區
        backup_section = SectionFrame(parent, title="備份與還原")
        backup_section.pack(fill="x", pady=(0, 20))
        
        ModernButton(backup_section.content, text="立即備份資料庫", icon='backup', style='success', command=self.backup_database).pack(anchor="w", pady=(0, 10))
        ModernButton(backup_section.content, text="從備份檔還原...", icon='refresh', style='danger', command=self.restore_database).pack(anchor="w")
        
        # 3. 分類管理
        cat_section = SectionFrame(parent, title="分類設定")
        cat_section.pack(fill="x", pady=(0, 20))
        ModernButton(cat_section.content, text="管理收支分類", icon='category', command=self.open_category_management).pack(anchor="w")
        
        # 4. 系統與說明
        sys_section = SectionFrame(parent, title="系統與說明")
        sys_section.pack(fill="x")
        
        h_sys = ctk.CTkFrame(sys_section.content, fg_color="transparent")
        h_sys.pack(fill="x")
        
        ModernButton(h_sys, text="快捷鍵說明 (F1)", icon='info', style='secondary', command=self.show_shortcuts_help).pack(side="left", padx=(0, 10))
        ModernButton(h_sys, text="關於本軟體", icon='info', style='secondary', command=self.show_about).pack(side="left", padx=(0, 10))
        ModernButton(h_sys, text="重新整理 (F5)", icon='refresh', style='secondary', command=self.refresh_data).pack(side="left")

    def _setup_view_report_embed(self, parent, report_type):
        """內嵌報表視圖：直接在頁面顯示圖表"""
        # 儲存當前報表類型與 parent 以供切換使用
        self.current_report_type = report_type
        self.current_report_parent = parent
        
        # 報表名稱對應 (簡化標題)
        titles = {
            "year_category": "年分類",
            "month_category": "月分類",
            "month_income_expense": "月收支",
            "daily_income_expense": "日收支"
        }
        title = titles.get(report_type, "報表分析")
        
        # Header
        header = ctk.CTkFrame(parent, fg_color="transparent")
        header.pack(fill="x", pady=(0, 15))
        
        ctk.CTkLabel(header, text=title, font=(FONTS['title'][0], 24, "bold")).pack(side="left")
        
        # 時間選擇區
        control_frame = ctk.CTkFrame(parent, fg_color=COLORS['bg_card'], corner_radius=10)
        control_frame.pack(fill="x", pady=(0, 15))
        
        control_content = ctk.CTkFrame(control_frame, fg_color="transparent")
        control_content.pack(fill="x", padx=15, pady=10)
        
        now = datetime.now()
        
        # 年份選擇 (儲存到 parent)
        ctk.CTkLabel(control_content, text="年份：").pack(side="left", padx=(0, 5))
        parent.year_var = tk.StringVar(value=str(now.year))
        year_combo = ctk.CTkComboBox(control_content, variable=parent.year_var, width=80,
                                     values=[str(y) for y in range(now.year - 5, now.year + 2)],
                                     command=lambda _: self._refresh_current_chart())
        year_combo.pack(side="left", padx=(0, 15))
        
        # 月份選擇 (儲存到 parent)
        parent.month_var = tk.StringVar(value=str(now.month))
        if report_type in ["month_category", "daily_income_expense"]:
            ctk.CTkLabel(control_content, text="月份：").pack(side="left", padx=(0, 5))
            month_combo = ctk.CTkComboBox(control_content, variable=parent.month_var, width=60,
                                          values=[str(m) for m in range(1, 13)],
                                          command=lambda _: self._refresh_current_chart())
            month_combo.pack(side="left", padx=(0, 15))
        
        # 圖表顯示區域 (儲存到 parent 物件屬性)
        chart_frame = tk.Frame(parent, bg=COLORS['bg_card'])
        chart_frame.pack(fill="both", expand=True)
        parent.chart_frame = chart_frame
        
        # 初始顯示圖表
        self.root.after(100, self._refresh_current_chart)
    
    def _refresh_current_chart(self):
        """刷新當前報表圖表"""
        if not hasattr(self, 'current_report_type') or not hasattr(self, 'current_report_parent'):
            return
        
        parent = self.current_report_parent
        if not hasattr(parent, 'chart_frame'):
            return
            
        chart_frame = parent.chart_frame
        
        # 清除舊圖表
        for widget in chart_frame.winfo_children():
            widget.destroy()
        
        try:
            # 從 parent 物件讀取年月變數
            year = int(parent.year_var.get()) if hasattr(parent, 'year_var') else datetime.now().year
            month = int(parent.month_var.get()) if hasattr(parent, 'month_var') else datetime.now().month
            report_type = self.current_report_type
            
            if report_type == "year_category":
                self.chart_manager.show_year_category_chart(chart_frame, year)
            elif report_type == "month_category":
                self.chart_manager.show_month_category_chart(chart_frame, year, month)
            elif report_type == "month_income_expense":
                self.chart_manager.show_month_income_expense_chart(chart_frame, year)
            elif report_type == "daily_income_expense":
                self.chart_manager.show_daily_income_expense_chart(chart_frame, year, month)
        except Exception as e:
            error_label = ctk.CTkLabel(chart_frame, text=f"圖表生成失敗：{e}", 
                                       text_color=COLORS['danger'])
            error_label.pack(expand=True)

    def on_transaction_select(self, event):
        """當選擇交易時顯示操作按鈕"""
        selected = self.transaction_tree.selection()
        if selected:
            # 顯示操作區域
            self.action_frame.pack(fill="x", padx=10, pady=(0, 10)) # 使用 pack 顯示
            
            # 取得選中的交易資訊
            item = selected[0]
            values = self.transaction_tree.item(item)['values']
            if values:
                date, trans_type, category, amount, *_ = values
                info_text = f"已選擇：{date} | {trans_type} | {category} | {amount}"
                self.selected_info_label.configure(text=info_text)
        else:
            self.action_frame.pack_forget()

    
        
        # 標題
        title_label = tk.Label(
            report_frame,
            text=f"{ICONS['chart']} 統計報表",
            font=FONTS['heading'],
            fg=COLORS['text_primary'],
            bg=COLORS['bg_primary']
        )
        title_label.pack(anchor='w', pady=(0, SPACING['md']))
        
        # 說明和按鈕容器
        content_frame = tk.Frame(
            report_frame,
            bg=COLORS['bg_card'],
            relief='solid',
            borderwidth=1
        )
        content_frame.pack(fill=tk.X)
        
        inner_frame = tk.Frame(content_frame, bg=COLORS['bg_card'])
        inner_frame.pack(fill=tk.X, padx=PADDING['loose'], pady=PADDING['loose'])
        
        # 說明文字
        desc_label = tk.Label(
            inner_frame,
            text="查看詳細的統計報表，包含年度分類、月度分類、月度收支、每日收支等圖表分析",
            font=FONTS['body'],
            fg=COLORS['text_secondary'],
            bg=COLORS['bg_card']
        )
        desc_label.pack(side=tk.LEFT, padx=(0, SPACING['lg']))
        
        # 開啟報表按鈕
        from .ui_components import ModernButton
        ModernButton(
            inner_frame,
            text="開啟報表視窗",
            style='primary',
            icon='chart',
            command=self.open_report_window
        ).pack(side=tk.RIGHT)
    
    # 篩選相關方法
    def on_filter_applied(self, filters: dict):
        """當篩選條件套用時的回調"""
        transactions = self.transaction_manager.get_transactions(limit=1000)
        
        filtered_transactions = []
        for trans in transactions:
            # 日期篩選
            if filters['start_date'] and trans['date'] < filters['start_date']:
                continue
            if filters['end_date'] and trans['date'] > filters['end_date']:
                continue
            
            # 類型篩選
            if filters['type'] != "all" and trans['type'] != filters['type']:
                continue
            
            # 分類篩選
            if filters['category'] != "全部分類" and filters['category'] and trans['category_name'] != filters['category']:
                continue
            
            # 關鍵字篩選
            if filters['keyword'] and filters['keyword'] not in str(trans.get('description', '')).lower():
                continue
            
            filtered_transactions.append(trans)
        
        self.display_transactions(filtered_transactions)
    
    def display_transactions(self, transactions):
        """顯示交易記錄"""
        # 防禦性檢查：確保 transaction_tree 已建立
        if not hasattr(self, 'transaction_tree'):
            return
        
        # 清除現有項目
        for item in self.transaction_tree.get_children():
            self.transaction_tree.delete(item)
        
        self.current_transactions = transactions
        
        for trans in transactions:
            # 移除小數點顯，改為千分位整數
            amount_display = f"${int(trans['amount']):,}"
            if trans['type'] == 'income':
                amount_display = f"+{amount_display}"
            else:
                amount_display = f"-{amount_display}"
            
            type_display = "收入" if trans['type'] == 'income' else "支出"
            tags = ('income' if trans['type'] == 'income' else 'expense',)
            
            self.transaction_tree.insert('', 'end', values=(
                trans['date'],
                type_display,
                trans['category_name'],
                amount_display,
                trans.get('description', '')
            ), tags=(str(trans['id']),) + tags)
        
        # 設定顏色
        self.transaction_tree.tag_configure('income', foreground='green')
        self.transaction_tree.tag_configure('expense', foreground='red')
        
        # 更新狀態
        if hasattr(self, 'list_status_label'):
             self.list_status_label.configure(text=f"共 {len(transactions)} 筆記錄")
        
        # 隱藏操作區域
        if hasattr(self, 'action_frame'):
            self.action_frame.pack_forget()
    
    def refresh_data(self):
        """重新整理資料顯示"""
        try:
            # 1. 刷新 Dashboard (統計數據)
            self.refresh_dashboard()
            
            # 2. 刷新 交易列表 (若存在)
            self.refresh_transactions()
            
            # 3. 刷新 分類篩選器 (若存在)
            if hasattr(self, 'filter_panel'):
                self.filter_panel.update_category_filter_options()
            
            self.status_label.configure(text="資料已更新")
            
        except Exception as e:
            messagebox.showerror("錯誤", f"資料更新失敗：{e}")
            if hasattr(self, 'status_label'):
                self.status_label.configure(text="更新失敗")

    def refresh_dashboard(self):
        """刷新 Dashboard 數據"""
        if hasattr(self, 'income_card'):
            self.update_statistics()
        # 重新載入交易列表 (根據當前選擇的期間)
        if hasattr(self, 'current_period'):
            self.filter_by_period(self.current_period)
    
    def filter_by_period(self, period):
        """根據期間篩選交易"""
        self.current_period = period
        
        # 更新按鈕樣式
        if hasattr(self, 'period_buttons'):
            for key, btn in self.period_buttons.items():
                if key == period:
                    btn.configure(fg_color=COLORS['primary'], text_color="white")
                else:
                    btn.configure(fg_color="transparent", text_color=COLORS['text_primary'])
        
        # 計算日期範圍
        now = datetime.now()
        if period == "today":
            start_date = now.strftime('%Y-%m-%d')
            end_date = start_date
        elif period == "week":
            from datetime import timedelta
            week_start = now - timedelta(days=now.weekday())
            week_end = week_start + timedelta(days=6)
            start_date = week_start.strftime('%Y-%m-%d')
            end_date = week_end.strftime('%Y-%m-%d')
        elif period == "month":
            import calendar
            start_date = f"{now.year}-{now.month:02d}-01"
            last_day = calendar.monthrange(now.year, now.month)[1]
            end_date = f"{now.year}-{now.month:02d}-{last_day}"
        elif period == "year":
            start_date = f"{now.year}-01-01"
            end_date = f"{now.year}-12-31"
        else:
            start_date = None
            end_date = None
        
        # 篩選交易
        transactions = self.transaction_manager.get_transactions(limit=500)
        if start_date and end_date:
            filtered = [t for t in transactions if start_date <= t['date'] <= end_date]
        else:
            filtered = transactions
        
        self.display_transactions(filtered)

    def refresh_transactions(self):
        """刷新交易列表數據"""
        if hasattr(self, 'transaction_tree'):
            # 默認重新載入最新 200 筆
            transactions = self.transaction_manager.get_transactions(limit=200)
            self.display_transactions(transactions)
    
    def update_statistics(self):
        """更新統計顯示"""
        now = datetime.now()
        summary = self.transaction_manager.get_monthly_summary(now.year, now.month)
        
        # 更新卡片數值
        if hasattr(self, 'income_card'):
            self.income_card.set_value(summary['total_income'])
        if hasattr(self, 'expense_card'):
            self.expense_card.set_value(summary['total_expense'])
        if hasattr(self, 'balance_card'):
            self.balance_card.set_value(summary['balance'])
    
    # update_filtered_statistics 已移除 (不再需要)
    
    def open_report_window(self, report_type="year_category"):
        """開啟報表視窗"""
        from .report_window import ReportWindow
        ReportWindow(self.root, self.transaction_manager, initial_report_type=report_type)
    
    # 交易管理方法
    def add_transaction(self):
        """新增交易記錄"""
        dialog = TransactionDialog(self.root, self.category_manager, self.transaction_manager)
        self.root.wait_window(dialog.dialog)
        
        if dialog.result:
            data = dialog.result
            success = self.transaction_manager.add_transaction(
                date=data['date'],
                transaction_type=data['type'],
                category_id=data['category_id'],
                amount=data['amount'],
                description=data['description']
            )
            
            if success:
                messagebox.showinfo("成功", "交易記錄新增成功！")
                self.refresh_data()
                self.status_label.configure(text="新增記錄成功")
            else:
                messagebox.showerror("錯誤", "交易記錄新增失敗！")
                self.status_label.configure(text="新增記錄失敗")
    
    def edit_transaction(self):
        """編輯選中的交易記錄"""
        selected_item = self.transaction_tree.selection()
        if not selected_item:
            messagebox.showwarning("提醒", "請先選擇要編輯的交易記錄")
            return
        
        transaction_id = int(self.transaction_tree.item(selected_item[0])['tags'][0])
        
        transactions = self.transaction_manager.get_transactions()
        transaction_data = None
        for trans in transactions:
            if trans['id'] == transaction_id:
                transaction_data = trans
                break
        
        if not transaction_data:
            messagebox.showerror("錯誤", "找不到交易記錄")
            return
        
        dialog = TransactionDialog(self.root, self.category_manager, 
                                 self.transaction_manager, transaction_data)
        self.root.wait_window(dialog.dialog)
        
        if dialog.result:
            data = dialog.result
            success = self.transaction_manager.update_transaction(
                transaction_id=transaction_id,
                date=data['date'],
                transaction_type=data['type'],
                category_id=data['category_id'],
                amount=data['amount'],
                description=data['description']
            )
            
            if success:
                messagebox.showinfo("成功", "交易記錄更新成功！")
                self.refresh_data()
                self.status_label.configure(text="更新記錄成功")
            else:
                messagebox.showerror("錯誤", "交易記錄更新失敗！")
                self.status_label.configure(text="更新記錄失敗")
    
    def delete_transaction(self):
        """刪除選中的交易記錄"""
        selected_item = self.transaction_tree.selection()
        if not selected_item:
            messagebox.showwarning("提醒", "請先選擇要刪除的交易記錄")
            return
        
        if not messagebox.askyesno("確認", "確定要刪除這筆交易記錄嗎？\n此操作無法復原。"):
            return
        
        transaction_id = int(self.transaction_tree.item(selected_item[0])['tags'][0])
        
        success = self.transaction_manager.delete_transaction(transaction_id)
        
        if success:
            messagebox.showinfo("成功", "交易記錄刪除成功！")
            self.refresh_data()
            self.status_label.configure(text="刪除記錄成功")
        else:
            messagebox.showerror("錯誤", "交易記錄刪除失敗！")
            self.status_label.configure(text="刪除記錄失敗")
    
    # 分類管理
    def open_category_management(self):
        """開啟分類管理對話框"""
        dialog = CategoryManagementDialog(self.root, self.category_manager, self.transaction_manager)
        self.root.wait_window(dialog.dialog)
        
        # 重新整理資料以更新分類選項
        self.refresh_data()
    
    # 匯出功能
    def export_to_csv(self):
        """匯出資料到 CSV 檔案"""
        if not self.current_transactions:
            messagebox.showwarning("提醒", "沒有資料可匯出")
            return
        
        try:
            filename = filedialog.asksaveasfilename(
                title="匯出 CSV 檔案",
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
                initialfile=f"記帳資料_{datetime.now().strftime('%Y%m%d')}.csv"
            )
        except Exception as e:
            messagebox.showerror("選擇檔案錯誤", f"無法開啟存檔對話框：{str(e)}")
            return
        
        if not filename:
            return
        
        try:
            with open(filename, 'w', newline='', encoding='utf-8-sig') as csvfile:
                writer = csv.writer(csvfile)
                
                # 寫入標題
                writer.writerow(['日期', '類型', '分類', '金額', '備註'])
                
                # 寫入交易資料
                for trans in self.current_transactions:
                    type_display = "收入" if trans['type'] == 'income' else "支出"
                    writer.writerow([
                        trans['date'],
                        type_display,
                        trans['category_name'],
                        trans['amount'],
                        trans.get('description', '')
                    ])
                
                # 寫入統計摘要
                writer.writerow([])
                writer.writerow(['統計摘要'])
                
                total_income = sum(trans['amount'] for trans in self.current_transactions if trans['type'] == 'income')
                total_expense = sum(trans['amount'] for trans in self.current_transactions if trans['type'] == 'expense')
                balance = total_income - total_expense
                
                writer.writerow(['總收入', f'${total_income:.2f}'])
                writer.writerow(['總支出', f'${total_expense:.2f}'])
                writer.writerow(['結餘', f'${balance:.2f}'])
                
                # 寫入匯出資訊
                writer.writerow([])
                writer.writerow(['匯出資訊'])
                writer.writerow(['匯出時間', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
                writer.writerow(['記錄筆數', len(self.current_transactions)])
            
            messagebox.showinfo("成功", f"資料已成功匯出到：\n{filename}")
            self.status_label.configure(text="CSV 匯出成功")
            
        except Exception as e:
            messagebox.showerror("錯誤", f"匯出失敗：{str(e)}")
            self.status_label.configure(text="CSV 匯出失敗")
    
    def export_to_excel(self):
        """匯出資料到 Excel 檔案"""
        if not self.current_transactions:
            messagebox.showwarning("提醒", "沒有資料可匯出")
            return
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            result = messagebox.askyesno("缺少套件", 
                "Excel 匯出需要安裝 openpyxl 套件。\n\n" +
                "請在終端機執行：pip install openpyxl\n\n" +
                "現在要改用 CSV 格式匯出嗎？")
            if result:
                self.export_to_csv()
            return
        
        try:
            filename = filedialog.asksaveasfilename(
                title="匯出 Excel 檔案",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=f"記帳資料_{datetime.now().strftime('%Y%m%d')}.xlsx"
            )
        except Exception as e:
            messagebox.showerror("選擇檔案錯誤", f"無法開啟存檔對話框：{str(e)}")
            return
        
        if not filename:
            return
        
        try:
            wb = openpyxl.Workbook()
            ws_data = wb.active
            ws_data.title = "交易記錄"
            
            # 設定標題樣式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            headers = ['日期', '類型', '分類', '金額', '備註']
            for col, header in enumerate(headers, 1):
                cell = ws_data.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # 寫入交易資料
            for row, trans in enumerate(self.current_transactions, 2):
                ws_data.cell(row=row, column=1, value=trans['date'])
                ws_data.cell(row=row, column=2, value="收入" if trans['type'] == 'income' else "支出")
                ws_data.cell(row=row, column=3, value=trans['category_name'])
                ws_data.cell(row=row, column=4, value=trans['amount'])
                ws_data.cell(row=row, column=5, value=trans.get('description', ''))
            
            # 調整欄寬
            column_widths = [12, 8, 15, 12, 30]
            for col, width in enumerate(column_widths, 1):
                ws_data.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
            
            wb.save(filename)
            messagebox.showinfo("成功", f"Excel 檔案已成功匯出到：\n{filename}")
            self.status_label.configure(text="Excel 匯出成功")
            
        except Exception as e:
            messagebox.showerror("錯誤", f"Excel 匯出失敗：{str(e)}")
            self.status_label.configure(text="Excel 匯出失敗")
    
    # 備份和還原
    def backup_database(self):
        """備份資料庫"""
        if not self.backup_manager:
            messagebox.showerror("錯誤", "備份功能不可用\n請確認 utils/backup.py 存在")
            return
        
        # 執行備份
        success, message = self.backup_manager.backup_database()
        
        if success:
            # 取得檔案大小
            import os
            file_size = os.path.getsize(message)
            size_str = format_file_size(file_size)
            
            messagebox.showinfo("備份成功", 
                f"資料庫已成功備份！\n\n"
                f"備份檔案: {os.path.basename(message)}\n"
                f"檔案大小: {size_str}\n"
                f"位置: backup/")
            self.status_label.configure(text="資料庫備份成功")
        else:
            messagebox.showerror("備份失敗", message)
            self.status_label.configure(text="備份失敗")
    
    def restore_database(self):
        """還原資料庫"""
        if not self.backup_manager:
            messagebox.showerror("錯誤", "還原功能不可用")
            return
        
        # 列出可用的備份
        backups = self.backup_manager.list_backups()
        
        if not backups:
            messagebox.showwarning("提示", "沒有可用的備份檔案\n\n請先執行備份功能")
            return
        
        # 建立還原對話框
        restore_dialog = tk.Toplevel(self.root)
        restore_dialog.title("還原資料庫")
        restore_dialog.geometry("500x400")
        restore_dialog.transient(self.root)
        restore_dialog.grab_set()
        
        ttk.Label(restore_dialog, text="選擇要還原的備份", 
                  font=("Arial", 12, "bold")).pack(pady=10)
        
        # 備份列表
        list_frame = ttk.Frame(restore_dialog)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        columns = ('檔名', '大小', '建立時間')
        backup_tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=10)
        
        for col in columns:
            backup_tree.heading(col, text=col)
        
        backup_tree.column('檔名', width=200)
        backup_tree.column('大小', width=100)
        backup_tree.column('建立時間', width=150)
        
        # 填入備份資料
        for backup in backups:
            backup_tree.insert('', 'end', values=(
                backup['name'],
                format_file_size(backup['size']),
                backup['created_time'].strftime('%Y-%m-%d %H:%M:%S')
            ), tags=(backup['path'],))
        
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=backup_tree.yview)
        backup_tree.configure(yscrollcommand=scrollbar.set)
        
        backup_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 按鈕
        button_frame = ttk.Frame(restore_dialog)
        button_frame.pack(pady=10)
        
        def do_restore():
            selected = backup_tree.selection()
            if not selected:
                messagebox.showwarning("提示", "請選擇要還原的備份")
                return
            
            backup_path = backup_tree.item(selected[0])['tags'][0]
            
            if not messagebox.askyesno("確認還原", 
                "確定要從備份還原資料庫嗎？\n\n"
                "⚠️ 警告：當前資料庫將被覆蓋！\n"
                "（系統會自動備份當前資料庫）"):
                return
            
            success, msg = self.backup_manager.restore_database(backup_path)
            
            if success:
                messagebox.showinfo("還原成功", 
                    f"{msg}\n\n請重新啟動程式以載入還原的資料")
                restore_dialog.destroy()
                self.status_label.configure(text="資料庫已還原")
            else:
                messagebox.showerror("還原失敗", msg)
        
        ttk.Button(button_frame, text="還原", command=do_restore).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="取消", command=restore_dialog.destroy).pack(side=tk.LEFT, padx=5)
    
    # 說明功能
    def show_shortcuts_help(self):
        """顯示快捷鍵說明"""
        help_text = """
快捷鍵說明

Ctrl+N    新增交易
F5        重新整理
Ctrl+S    匯出 CSV
Ctrl+M    分類管理
Alt+F4    退出程式

提示：
- 點擊交易列表中的「編輯」或「刪除」進行操作
- 雙擊交易記錄也可以快速編輯
        """
        messagebox.showinfo("快捷鍵說明", help_text)
    
    def show_about(self):
        """顯示關於對話框"""
        about_text = """
個人記帳本 v1.1

重構版本 - 模組化設計

功能特色：
• 交易記錄管理
• 分類管理
• 進階篩選
• 統計報表
• 圖表分析
• 資料匯出

開發：Python + tkinter
        """
        messagebox.showinfo("關於", about_text)
    
    def on_closing(self):
        """程式關閉時的處理"""
        if messagebox.askokcancel("退出", "確定要退出個人記帳本嗎？"):
            self.root.destroy()
    
    def run(self):
        """啟動主程式"""
        # 綁定關閉事件
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        # 顯示啟動訊息
        self.status_label.configure(text="個人記帳本已啟動")
        
        print("🚀 個人記帳本已啟動 (重構版)")
        print("📚 使用說明：")
        print("   - Ctrl+N: 新增交易")
        print("   - Ctrl+E: 編輯交易")
        print("   - Del: 刪除交易")
        print("   - F5: 重新整理")
        print("   - Ctrl+S: 匯出 CSV")
        print("   - Ctrl+M: 分類管理")
        
        # 啟動主迴圈
        self.root.mainloop()
